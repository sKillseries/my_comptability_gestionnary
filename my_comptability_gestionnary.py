#!/usr/bin/python3

from openpyxl import Workbook, load_workbook
from termcolor import colored

def get_user_input(message):
    return input(message).strip()

def calculate_expenses_variables():
    print(colored("""
###########################
    Saisie des dépenses variables :
###########################
    """, 'magenta'))
    retrait = float(get_user_input("Combien d'argent avez-vous retiré au guichet ce mois-ci ? : "))
    coiffeur = float(get_user_input("Combien a coûté le coiffeur ? : "))
    vetement = float(get_user_input("Combien avez-vous dépensé en vêtements ce mois-ci ? : "))
    alimentation = float(get_user_input("A combien s'élève le montant de vos courses ce mois-ci ? : "))
    cosmetique = float(get_user_input("Combien avez-vous dépensé en cosmétique ? : "))
    loisir = float(get_user_input("Combien as-tu dépensé en loisir ? : "))
    autre = float(get_user_input("Montant autre dépense ? : "))
    depenses_variables = round(retrait + coiffeur + vetement + alimentation + cosmetique + loisir + autre, 2)
    print(colored(f"Vos dépenses variables s'élèvent à {depenses_variables}€ ce mois-ci.", 'yellow'))
    return depenses_variables

def calculate_expenses_fixes():
    print(colored("""
##################################
    Saisie des dépenses fixes :
##################################
    """, 'magenta'))
    epargne = float(get_user_input("Combien avez-vous épargné ce mois-ci ? : "))
    investissement = float(get_user_input("Combien avez-vous investi ce mois-ci ? : "))
    telephone = float(get_user_input("Combien coûte votre abonnement téléphonique ce mois-ci ? : "))
    internet = float(get_user_input("Combien coûte votre abonnement internet ce mois-ci ? : "))
    loyer = float(get_user_input("Combien coûte votre loyer ce mois-ci ? : "))
    transport = float(get_user_input("Combien coûte votre abonnement de transport ce mois-ci ? : "))
    auto = float(get_user_input("Combien coûte votre traite auto ce mois-ci ? : "))
    electricite = float(get_user_input("Quelle est la somme de votre facture d'électricité ce mois-ci ? : "))
    eau = float(get_user_input("Quelle est la somme de votre facture d'eau ce mois-ci ? : "))
    depenses_fixes = round(epargne + investissement + telephone + internet + loyer + transport + auto + electricite + eau, 2)
    print(colored(f"Vos dépenses fixes s'élèvent à {depenses_fixes}€ ce mois-ci.", 'yellow'))
    return depenses_fixes

def calculate_revenues():
    print(colored("""
######################################
    Saisie des revenus :
######################################
    """, 'magenta'))
    salaire = float(get_user_input("Quelle est la somme de votre salaire ce mois-ci ? : "))
    prime = float(get_user_input("Quelle a été la somme de votre prime ce mois-ci ? : "))
    locations = float(get_user_input("Quelle est la somme de vos revenus de location ce mois-ci ? : "))
    dividendes = float(get_user_input("Quelle est la somme de vos dividendes ce mois-ci ? : "))
    interets = float(get_user_input("Quelle est la somme de vos intérêts ce mois-ci ? : "))
    redevances = float(get_user_input("Quelle est la somme de vos redevances ce mois-ci ? : "))
    revenues = round(salaire + prime + locations + dividendes + interets + redevances, 2)
    print(colored(f"Vos revenus ce mois-ci s'élèvent à {revenues}€.", 'green'))
    return revenues

def save_and_proceed(revenues, depenses_fixes, depenses_variables):
    reste = round(revenues - depenses_fixes - depenses_variables, 2)
    print(colored(f"""
#####################################
    Bilan comptable de la saisie :
#####################################
Revenus : {revenues}€
Dépenses fixes : {depenses_fixes}€
Dépenses variables : {depenses_variables}€
Restes : {reste}€
    """, 'cyan'))
    print(colored("La saisie a été enregistrée dans le fichier my_comptability_sheet.xlsx.", 'green'))
    answer = get_user_input("Voulez-vous saisir un autre mois (o/N) ? : ")
    if answer.lower() in ["o", "y", "yes"]:
        menu()
    else:
        print("Fermeture du programme.")
        exit(0)

def menu():
    print(
        """
##############################
    Enregistrement saisie :
##############################

Quel mois voulez-vous enregistrer la saisie ?
> 1 : Janvier
> 2 : Février
> 3 : Mars
> 4 : Avril
> 5 : Mai
> 6 : Juin
> 7 : Juillet
> 8 : Août
> 9 : Septembre
> 10: Octobre
> 11: Novembre
> 12: Décembre
> q : Quitter
        """
    )
    mois = get_user_input("Veuillez sélectionner le numéro du mois : ")
    if mois.lower() == "q":
        print("Fermeture du programme.")
        exit(0)
    elif mois.isdigit() and 1 <= int(mois) <= 12:
        sheet.cell(row=int(mois) + 4, column=3).value = revenues
        sheet.cell(row=int(mois) + 4, column=4).value = depenses_fixes
        sheet.cell(row=int(mois) + 4, column=5).value = depenses_variables
        wb.save("my_comptability_sheet.xlsx")
        save_and_proceed(revenues, depenses_fixes, depenses_variables)
    else:
        print(colored("Veuillez sélectionner une option présente dans la liste !", 'yellow'))
        menu()

def main():
    global wb
    global sheet
    annee = get_user_input(colored("En quelle année sommes-nous ? ", 'blue'))
    wb = load_workbook("my_comptability_sheet.xlsx")
    sheet = wb[f'{annee}']
    global revenues
    global depenses_fixes
    global depenses_variables
    revenues = calculate_revenues()
    depenses_fixes = calculate_expenses_fixes()
    depenses_variables = calculate_expenses_variables()
    menu()

if __name__ == '__main__':
    main()